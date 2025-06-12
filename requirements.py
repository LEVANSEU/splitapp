import streamlit as st
import pandas as pd
import io
from openpyxl import Workbook
import re

st.set_page_config(layout="wide")
st.markdown("""
    <style>
        body, .main, .block-container {
            background-color: white !important;
            color: #222 !important; /* Default text color */
            font-family: 'Segoe UI', sans-serif;
        }
        h1, h2, h3, h4, h5, h6, .stMarkdown, .stText, .stTextLabelWrapper, label {
            color: #222 !important;
        }
        .stFileUploader, .stTextInput, .stSelectbox, .stRadio, .stButton, .stDataFrame,
        .stTextInput input, .stSelectbox div[data-baseweb="select"],
        .stSelectbox div[data-baseweb="select"] *,
        .stRadio div[role="radiogroup"] label,
        .stRadio div[role="radiogroup"] label * {
            background-color: #f5f5f5 !important;
            color: #222 !important;
            border-radius: 10px;
            font-size: 14px !important;
        }
        .stFileUploader {
            max-width: 600px !important;
            margin: 0 auto !important;
        }
        .stButton>button {
            background-color: #4CAF50;
            color: white !important;
            font-weight: bold;
            border: none;
            border-radius: 8px;
            padding: 6px 14px;
            font-size: 14px;
        }
        .stButton>button:hover {
            background-color: #45a049;
        }
        .summary-header {
            display: flex;
            font-weight: bold;
            margin-top: 1em;
            padding-bottom: 0.5rem;
            border-bottom: 2px solid #999;
            text-align: center;
            background-color: #f0f0f0;
            border-radius: 8px;
            color: #222 !important;
        }
        .summary-header div {
            flex: 1;
            padding: 0.5rem;
        }
        .number-cell {
            text-align: right !important;
            font-variant-numeric: tabular-nums;
            padding-right: 1rem;
            font-weight: bold;
            color: #222;
        }
        /* Ensure table text is visible and numbers aligned right on white background */
        .stTable {
            color: #222 !important;
        }
        .stTable td, .stTable th {
            color: #222 !important;
            background-color: white !important;
        }
        .stTable td:nth-child(n+3), .stTable th:nth-child(n+3) { /* Apply to numeric columns (3rd and beyond) */
            text-align: right !important;
            font-variant-numeric: tabular-nums;
            padding-right: 1rem;
        }
        /* Ensure write output and other elements are visible */
        .stMarkdown, .stWrite {
            color: #222 !important;
        }
        .dataFrame {
            color: #222 !important;
        }
    </style>
""", unsafe_allow_html=True)

st.title("Excel გენერატორი")

report_file = st.file_uploader("ატვირთე ანგარიშფაქტურების ფაილი (report.xlsx)", type=["xlsx"])
statement_files = st.file_uploader("ატვირთე საბანკო ამონაწერის ფაილები (statement.xlsx)", type=["xlsx"], accept_multiple_files=True)

if report_file and statement_files:
    try:
        st.write(f"Uploaded report file: {report_file.name}")
        purchases_df = pd.read_excel(report_file, sheet_name='Grid')
        st.write("purchases_df head:", purchases_df.head())
    except Exception as e:
        st.error(f"Error reading report file: {str(e)}")
        purchases_df = pd.DataFrame()

    # Process multiple bank statement files
    bank_dfs = []
    if statement_files:
        for statement_file in statement_files:
            try:
                st.write(f"Processing statement file: {statement_file.name}")
                df = pd.read_excel(statement_file)
                st.write(f"Raw df head for {statement_file.name}:", df.head())
                df['P'] = df.iloc[:, 15].astype(str).str.strip()  # Identification code
                df['Name'] = df.iloc[:, 14].astype(str).str.strip()  # Name from column O
                df['Amount'] = pd.to_numeric(df.iloc[:, 3], errors='coerce').fillna(0)  # Amount from column D
                bank_dfs.append(df)
            except Exception as e:
                st.error(f"Error reading statement file {statement_file.name}: {str(e)}")
                continue
    
    # Combine all bank statement DataFrames
    bank_df = pd.concat(bank_dfs, ignore_index=True) if bank_dfs else pd.DataFrame()
    st.write("bank_df head:", bank_df.head())

    purchases_df['დასახელება'] = purchases_df['გამყიდველი'].astype(str).apply(lambda x: re.sub(r'^\(\d+\)\s*', '', x).strip())
    purchases_df['საიდენტიფიკაციო კოდი'] = purchases_df['გამყიდველი'].apply(lambda x: ''.join(re.findall(r'\d', str(x)))[:11])

    wb = Workbook()
    wb.remove(wb.active)

    ws1 = wb.create_sheet(title="ანგარიშფაქტურები კომპანიით")
    ws1.append(['დასახელება', 'საიდენტიფიკაციო კოდი', 'ანგარიშფაქტურების ჯამი', 'ჩარიცხული თანხა', 'სხვაობა'])

    company_summaries = []

    for company_id, group in purchases_df.groupby('საიდენტიფიკაციო კოდი'):
        try:
            company_name = group['დასახელება'].iloc[0]
            unique_invoices = group.groupby('სერია №')['ღირებულება დღგ და აქციზის ჩათვლით'].sum().reset_index()
            company_invoice_sum = unique_invoices['ღირებულება დღგ და აქციზის ჩათვლით'].sum()

            paid_sum = bank_df[bank_df["P"] == str(company_id)]['Amount'].sum()
            difference = company_invoice_sum - paid_sum

            ws1.append([company_name, company_id, company_invoice_sum, paid_sum, difference])
            company_summaries.append((company_name, company_id, company_invoice_sum, paid_sum, difference))
        except Exception as e:
            st.error(f"Error processing company {company_id}: {str(e)}")
            continue

    st.write("company_summaries length:", len(company_summaries))  # Diagnostic for summaries

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    if 'selected_company' not in st.session_state:
        st.subheader("📋 კომპანიების ყჩამონათვალი")  # Note: "ჩ" seems to be a typo, likely meant "ჩამონათვალი"

        search_code = st.text_input("🔎 ჩაწერე საიდენტიფიკაციო კოდი:", "")
        sort_column = st.selectbox("📊 დალაგების ველი", ["ინვოისების ჯამი", "ჩარიცხვა", "სხვაობა"])
        sort_order = st.radio("⬆️⬇️ დალაგების ტიპი", ["ზრდადობით", "კლებადობით"])

        sort_index = {"ინვოისების ჯამი": 2, "ჩარიცხვა": 3, "სხვაობა": 4}[sort_column]
        reverse = sort_order == "კლებადობით"

        filtered_summaries = company_summaries
        if search_code.strip():
            filtered_summaries = [item for item in company_summaries if item[1] == search_code.strip()]

        filtered_summaries = sorted(filtered_summaries, key=lambda x: x[sort_index], reverse=reverse)

        st.markdown("""
        <div class='summary-header'>
            <div style='flex: 2;'>დასახელება</div>
            <div style='flex: 2;'>საიდენტიფიკაციო კოდი</div>
            <div style='flex: 1.5;'>ინვოისების ჯამი</div>
            <div style='flex: 1.5;'>ჩარიცხვა</div>
            <div style='flex: 1.5;'>სხვაობა</div>
        </div>
        """, unsafe_allow_html=True)

        for name, company_id, invoice_sum, paid_sum, difference in filtered_summaries:
            col1, col2, col3, col4, col5 = st.columns([2, 2, 1.5, 1.5, 1.5])
            with col1:
                st.markdown(name)
            with col2:
                if st.button(f"{company_id}", key=f"id_{company_id}"):
                    st.session_state['selected_company'] = company_id
                    st.write(f"Selected company: {company_id}")  # Debug
            with col3:
                st.markdown(f"<div class='number-cell'>{invoice_sum:,.2f}</div>", unsafe_allow_html=True)
            with col4:
                st.markdown(f"<div class='number-cell'>{paid_sum:,.2f}</div>", unsafe_allow_html=True)
            with col5:
                st.markdown(f"<div class='number-cell'>{difference:,.2f}</div>", unsafe_allow_html=True)

        # Sorting and filtering options
        if 'sort_order_missing' not in st.session_state:
            st.session_state['sort_order_missing'] = "კლებადობით"
        if 'search_query_missing' not in st.session_state:
            st.session_state['search_query_missing'] = ""

        sort_order_missing = st.radio("სორტირება:", ["ზრდადობით", "კლებადობით"], key="sort_order_missing")
        search_query_missing = st.text_input("ძებნა (კოდი ან დასახელება):", key="search_query_missing")

        # New button with logic for companies not in invoice list
        if st.button("დამატებითი მოქმედება"):
            # Get unique company IDs from bank_df
            bank_company_ids = bank_df['P'].unique()
            # Get company IDs from purchases_df
            invoice_company_ids = purchases_df['საიდენტიფიკაციო კოდი'].unique()
            # Find companies in bank_df but not in purchases_df
            missing_company_ids = [cid for cid in bank_company_ids if cid not in invoice_company_ids]
            
            if missing_company_ids:
                st.subheader("კომპანიები ანგარიშფაქტურის სიაში არ არიან")
                missing_data = []
                for company_id in missing_company_ids:
                    # Get company name from bank_df where P matches
                    matching_rows = bank_df[bank_df['P'] == str(company_id)]
                    company_name = matching_rows['Name'].iloc[0] if not matching_rows.empty else "-"
                    total_amount = bank_df[bank_df['P'] == str(company_id)]['Amount'].sum()
                    invoice_amount = 0.00  # Since they are not in invoice list
                    difference = total_amount - invoice_amount
                    missing_data.append([company_name, company_id, total_amount, invoice_amount, difference])
                
                # Apply search filter
                if search_query_missing.strip():
                    missing_data = [item for item in missing_data if 
                                  str(item[1]) == search_query_missing.strip() or 
                                  str(item[0]).lower().find(search_query_missing.lower().strip()) != -1]
                
                # Apply sort
                sort_reverse = st.session_state['sort_order_missing'] == "კლებადობით"
                missing_data.sort(key=lambda x: x[2], reverse=sort_reverse)  # Sort by total amount
                
                # Display as a table with headers
                st.markdown("""
                <div class='summary-header'>
                    <div style='flex: 2;'>დასახელება</div>
                    <div style='flex: 2;'>საიდენტიფიკაციო კოდი</div>
                    <div style='flex: 1.5;'>ჩარიცხული თანხა</div>
                    <div style='flex: 1.5;'>ანგარიშფაქტურის თანხა</div>
                    <div style='flex: 1.5;'>სხვაობა</div>
                </div>
                """, unsafe_allow_html=True)
                
                for item in missing_data:
                    col1, col2, col3, col4, col5 = st.columns([2, 2, 1.5, 1.5, 1.5])
                    with col1:
                        st.write(item[0])
                    with col2:
                        if st.button(str(item[1]), key=f"missing_{item[1]}"):
                            st.session_state['selected_missing_company'] = item[1]  # Save selected company ID
                            st.write(f"Selected missing company: {item[1]}")  # Debug
                    with col3:
                        st.write(f"{item[2]:,.2f}")
                    with col4:
                        st.write(f"{item[3]:,.2f}")
                    with col5:
                        st.write(f"{item[4]:,.2f}")
            else:
                st.info("ყველა კომპანია ანგარიშფაქტურის სიაში გამოჩნდა.")

    else:
        selected_code = st.session_state['selected_company']
        df_full = pd.read_excel(report_file, sheet_name='Grid')
        df_full['დასახელება'] = df_full['გამყიდველი'].astype(str).apply(lambda x: re.sub(r'^\(\d+\)\s*', '', x).strip())
        df_full['საიდენტიფიკაციო კოდი'] = df_full['გამყიდველი'].apply(lambda x: ''.join(re.findall(r'\d', str(x)))[:11])
        matching_df = df_full[df_full['საიდენტიფიკაციო კოდი'] == selected_code]

        if not matching_df.empty:
            company_name = matching_df['დასახელება'].iloc[0]
            st.subheader(f"🔎 ({selected_code}) {company_name} - ანგარიშფაქტურები")
            st.dataframe(matching_df, use_container_width=True)

            st.subheader("🔍 მოძებნე გუგლში მასალა ან მომსახურება")
            col1, col2 = st.columns([3, 1])
            with col1:
                search_term = st.text_input("ჩაწერე სახელი ან სიტყვა:")
            with col2:
                if st.button("ძებნა"):
                    if search_term.strip():
                        search_url = f"https://www.google.com/search?q={search_term.replace(' ', '+')}"
                        st.markdown(f"[🌐 გადადი გუგლზე]({search_url})", unsafe_allow_html=True)
                    else:
                        st.warning("გთხოვ ჩაწერე текста ძებნამდე.")

            company_output = io.BytesIO()
            company_wb = Workbook()
            ws = company_wb.active
            ws.title = company_name[:31]
            ws.append(matching_df.columns.tolist())
            for row in matching_df.itertuples(index=False):
                ws.append(row)
            company_wb.save(company_output)
            company_output.seek(0)

            st.download_button(
                label=f"⬇️ ჩაღონათვირთე {company_name} ინვოისების Excel",
                data=company_output,
                file_name=f"{company_name}_ინვოისები.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        else:
            st.warning("📭 ჩანაწერი ვერ მოიძებნა ამ კომპანიისთვის.")

        if st.button("⬅️ დაბრუნება სრულ სიაზე"):
            del st.session_state['selected_company']

    # Detail view for selected missing company
    if 'selected_missing_company' in st.session_state:
        selected_id = st.session_state['selected_missing_company']
        st.subheader(f"ჩარიცხვების ცხრილი - {selected_id}")
        matching_transactions = bank_df[bank_df['P'] == str(selected_id)]
        if not matching_transactions.empty:
            st.table(matching_transactions[['Name', 'P', 'Amount']])  # Display transactions table
        else:
            st.warning("ჩანაწერი არ მოიძებნა ამ კომპანიისთვის.")
        if st.button("⬅️ დაბრუნება სრულ სიაზე"):
            del st.session_state['selected_missing_company']
            st.experimental_rerun()
